import cv2
import numpy as np
import openpyxl
import pandas as pd
from PIL import Image
import streamlit as st

recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read('detector/trainingData.yml')

workbook = openpyxl.load_workbook('ExamSeatArrangement.xlsx')
sheet = workbook.active

# Load ROIs coordinates from roi.csv
roi_df = pd.read_csv('roi.csv')

def mark_attendance(image):
    gray_image = cv2.cvtColor(image, cv2.COLOR_RGB2GRAY)

    attendance = []

    # Initialize an empty dictionary
    names = {}

    # Start from the second row (assuming the first row is the header)
    row_number = 2

    # Loop until there are no entries in the second column
    while sheet.cell(row=row_number, column=1).value is not None:
        # Get the value from the second column
        cell_value = sheet.cell(row=row_number, column=1).value
        sheet.cell(row=row_number, column=3).value = "Absent"
        
        # Add the value to the dictionary with the key as the current row number
        names[row_number - 1] = cell_value
        
        # Move to the next row
        row_number += 1

    for (x, y, w, h) in faces:
        face = gray_image[y:y+h, x:x+w]

        id_, confidence = recognizer.predict(face)
        print(id_, confidence)

        if confidence < 130:
            name = names.get(id_, "Unknown")
        else:
            name = "Unknown"
        attendance.append(name)

    present_count = 0
    for name in attendance:
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == name:
                sheet.cell(row=row, column=3).value = "Present"
                present_count += 1
                break

    workbook.save('ExamSeatArrangement.xlsx')
    return present_count

def capture_image():
    cap = cv2.VideoCapture(0)
    
    if not cap.isOpened():
        st.error("Error: Unable to open camera.")
        return
    
    cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
    
    while True:
        ret, frame = cap.read()
        
        if not ret:
            st.warning("Warning: Unable to capture frame from camera.")
            break
        
        cv2.imshow('Capture', frame)
        
        if cv2.waitKey(1) & 0xFF == ord(' '):
            cv2.imwrite('captured_image.png', frame)
            break
    
    cap.release()
    cv2.destroyAllWindows()

# Define a function to apply conditional formatting
def color_cell(value):
    if value == 'Present':
        color = 'green'
    elif value == 'Absent':
        color = 'red'
    else:
        color = 'black'  # default color
    return f'color: {color}'

# Streamlit app
def main():
    st.title("Exam Attendance System")

    option = st.radio("Select Option:", ("Upload Image", "Capture Image", "View Attendance Sheet"))

    if option == "Upload Image":
        uploaded_file = st.file_uploader("Upload Image", type=["jpg", "jpeg", "png"])
        if uploaded_file is not None:
            image = Image.open(uploaded_file)
            st.image(image, caption='Uploaded Image', use_column_width=True)

            if st.button("Mark Attendance"):
                count = mark_attendance(np.array(image))
                st.success(f"Attendance Marked: {count} Students Identified")

    elif option == "Capture Image":
        st.write("Click the button below to capture an image from the camera and mark attendence")
        if st.button("Capture"):
            capture_image()
            captured_image = Image.open('captured_image.png')
            st.image(captured_image, caption='Captured Image', use_column_width=True)

            mark_attendance(np.array(captured_image))
            st.success("Attendance Marked!")

    elif option == "View Attendance Sheet":
        # df = pd.read_excel('ExamSeatArrangement.xlsx')
        # st.write(df)

        # Read the Excel file into a DataFrame
        df = pd.read_excel('ExamSeatArrangement.xlsx')

        # Apply conditional formatting to the DataFrame
        styled_df = df.style.applymap(color_cell, subset=['ATT. STATUS'])

        # Display the styled DataFrame in Streamlit
        st.dataframe(styled_df)


if __name__ == "__main__":
    main()