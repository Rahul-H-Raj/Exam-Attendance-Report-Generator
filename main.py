import json
import threading
from flask import Blueprint, render_template, flash, redirect, url_for, request, jsonify, send_file
from io import BytesIO
from flask_login import login_required, current_user
from __init__ import create_app, db
import time
from models import Classroom
import cv2
import os
import pandas as pd
import openpyxl
from bs4 import BeautifulSoup
from report_generator import create_attendance_report
from excel_data_reader import extract_excel_to_html
from save_rois import save_regions
from normalize_regions import extract_and_save_regions
from show_regions import display_image_with_regions
from value_thresholds import set_threshold
from trainner import train

# our main blueprint
main = Blueprint('main', __name__)

# Path to your video file
VIDEO_PATH = 'Classrooms/A 201/During-Exam.mp4'
# Folder to save extracted frames
FRAMES_FOLDER = 'static'
# Constants for threshold values
THRESHOLDS = {
    1: 50, 2: 60, 3: 60, 4: 60, 
    5: 85, 6: 89, 7: 85, 8: 80, 
    9: 100, 10: 100, 11: 110, 12: 110
}

# Function to extract frames from the video at regular intervals
def extract_frames(interval_sec, no_of_frames):
    cap = cv2.VideoCapture(VIDEO_PATH)
    frame_rate = cap.get(cv2.CAP_PROP_FPS)
    interval_frames = int(interval_sec * frame_rate)
    current_frame = 1
    n = 0

    while cap.isOpened():
        ret, frame = cap.read()
        if not ret:
            break

        if current_frame % interval_frames == 0:
            n += 1
            # Save the frame as an image
            image_path = os.path.join(FRAMES_FOLDER, f'frame_{n}.jpg')
            cv2.imwrite(image_path, frame)
            yield image_path, time.strftime('%Y-%m-%d %H:%M:%S'), n

        if n == no_of_frames:
            break

        current_frame += 1

    cap.release()

def clear_folder():
    # Get a list of all files in the folder
    files = os.listdir(FRAMES_FOLDER)
    
    # Loop through each file and delete if it starts with "frames" and ends with an image extension
    for file in files:
        if file.startswith("frame") and file.endswith(".jpg"):
            os.remove(os.path.join(FRAMES_FOLDER, file))

# Function to predict on regions using LBPH model
def detect_presence(image, roi_df, lbph_model):
    predictions = {}
    for index, row in roi_df.iterrows():
        x, y, w, h = row['x'], row['y'], row['width'], row['height']
        # Extract region from the image
        region_img = image[y:y+h, x:x+w]
        # Convert region image to grayscale
        gray_region = cv2.cvtColor(region_img, cv2.COLOR_BGR2GRAY)
        # Make prediction using LBPH model
        label, confidence = lbph_model.predict(gray_region)
        predictions.update({index+1: (label, confidence)})

    print(predictions)
    return predictions

def mark_attendance(predictions, classroom_no):
    try:
        # Load the Excel workbook
        workbook_path = 'ExamSeatArrangement.xlsx'
        workbook = openpyxl.load_workbook(workbook_path)
        sheet = workbook[classroom_no]

        # Loop through rows
        for row in range(2, sheet.max_row + 1):
            if all(sheet.cell(row=row, column=col).value is None for col in range(1, 5)):
                break
            for col in range(1, sheet.max_column + 1, 4):
                sheet.cell(row=row, column=col + 3, value="Present")
                seat_no = sheet.cell(row=row, column=col).value
                if seat_no in predictions:
                    prediction = predictions[seat_no]
                    if prediction[1] < THRESHOLDS.get(seat_no, 0):
                        sheet.cell(row=row, column=col + 3, value="Absent")

        # Save the changes to the workbook
        workbook.save(workbook_path)
    except Exception as e:
        print(f"An error occurred while marking attendance: {str(e)}")

def format_numbers(numbers):
    if len(numbers) == 1:
        return 'bench number: ' + str(numbers[0]) + " is "
    elif len(numbers) == 2:
        return f"bench numbers: {numbers[0]} and {numbers[1]}" + " are "
    else:
        return f"bench numbers: {', '.join(map(str, numbers[:-1]))}, and {numbers[-1]} are "

def verify_check(predictions, classroom_no, thresholds):
    bench_ids = []
    # Loop through rows
    for key in predictions:
        prediction = predictions[key]
        if prediction[1] > thresholds.get(str(key)):
            bench_ids.append(key)
        
    if not bench_ids:
        return "Consistency check successful for classroom {}".format(classroom_no)
    
    str_format = format_numbers(bench_ids)
    return str_format + 'not properly detected. Please verify if each bench is present and correctly aligned, and redefine the regions for those IDs if necessary.'

@main.route('/consistency_check', methods=['POST'])
def consistency_check():
    if request.method == 'POST':
        classroom_no = request.json.get('classroom_no')
        classroom = Classroom.query.filter_by(classroom_no=classroom_no).first()
        if not classroom.seat_distribution:
            result = "The setup for Classroom {} is still pending, so the seating arrangement has not been determined and saved yet.".format(classroom_no)
        else:
            json_data = json.loads(classroom.seat_distribution)

            # Load saved LBPH model
            lbph_model_path = f'Classrooms/{classroom_no}/detector/trainingData.yml'
            lbph_model = cv2.face.LBPHFaceRecognizer_create()
            lbph_model.read(lbph_model_path)

            # Load ROIs coordinates from roi.csv
            roi_csv_path = f'Classrooms/{classroom_no}/roi.csv'
            roi_df = pd.read_csv(roi_csv_path)
            image = cv2.imread('frame_0.jpg')

            pred = detect_presence(image, roi_df, lbph_model)
            result = verify_check(pred, classroom_no, json_data)
        return result

@main.route('/define_rois', methods=['POST'])
def define_rois():
    data = request.get_json()
    classroom_no = data.get('classroom_no')
    bench_number = data.get('bench_number')

    image_path = 'frame_1.jpg'
    # Load existing region coordinates from roi.csv
    csv_path = f'Classrooms/{classroom_no}/regions.csv'
    if not os.path.exists(csv_path):
        return jsonify({'message': 'ROI file not found', 'success': False}), 404
    
    save_regions(image_path, bench_number, csv_path)
    extract_and_save_regions(image_path, csv_path)

    response = {
        'status': 'success',
        'message': f'ROIs defined for classroom {classroom_no}'
    }
    return jsonify(response)

@main.route('/show_regions', methods=['POST'])
def show_regions():
    data = request.get_json()
    classroom_no = data.get('classroom_no')

    image_path = 'frame_1.jpg'
    csv_path = f'Classrooms/{classroom_no}/roi.csv'

    display_image_with_regions(image_path, csv_path)

    response = {
        'status': 'success',
        'message': f'Regions shown for classroom {classroom_no}'
    }
    return jsonify(response)

@main.route('/') # home page that return 'index'
def index():
    return render_template('index.html')

@main.route('/profile') # profile page that return 'profile'
@login_required
def profile():
    return render_template('profile.html', name=current_user.name)

@main.route('/home')
def home():
    # Fetch classrooms data
    classrooms = Classroom.query.all()
    return render_template('home.html', classrooms=classrooms)  # Pass classrooms data to the template

# Route for delayed redirection from profile to home
@main.route('/redirect-home')
@login_required
def redirect_home():
    time.sleep(1)  # Delay for one second
    return redirect(url_for('main.home'))

@main.route('/save_classroom', methods=['POST'])
def save_classroom():
    if request.method == 'POST':
        data = request.json
        classroom_no = data.get('classroom_no')
        # Insert the classroom_no into the database (assuming Classroom is the model)
        classroom = Classroom(classroom_no=classroom_no)
        db.session.add(classroom)
        db.session.commit()
        return jsonify({'message': 'Inserted successfully'}), 200
    else:
        return jsonify({'error': 'Invalid request method'}), 405  # Method Not Allowed
    
@main.route('/classroom/<classroom_no>')
def classroom(classroom_no):
    clear_folder()
    # You can use the classroom_no parameter to fetch additional data or perform any other operations
    return render_template('classroom.html', classroom_no=classroom_no)

@main.route('/extract_frames', methods=['POST'])
def extract_frames_route():
    data = request.get_json()
    interval_sec = int(data.get('interval'))
    no_of_frames = int(data.get('no_of_frames'))
    extracted_frames = []

    for image_path, extraction_time, index in extract_frames(interval_sec, no_of_frames):
        extracted_frames.append({'image_path': image_path, 'extraction_time': extraction_time, 'index': index})
    time.sleep(4)
    return jsonify(extracted_frames), 200

@main.route('/trainner/<classroom_no>')
def trainner(classroom_no):
    return render_template('trainner.html', classroom_no=classroom_no)

@main.route('/start-training/<classroom_no>', methods=['POST'])
def start_training(classroom_no):
    path = f'Classrooms/{classroom_no}/dataset'
    train(path)
    # Load saved LBPH model
    lbph_model_path = f'Classrooms/{classroom_no}/detector/trainingData.yml'
    lbph_model = cv2.face.LBPHFaceRecognizer_create()
    lbph_model.read(lbph_model_path)

    # Load ROIs coordinates from roi.csv
    roi_csv_path = f'Classrooms/{classroom_no}/roi.csv'
    roi_df = pd.read_csv(roi_csv_path)
    image = cv2.imread('frame_1.jpg')

    pred = detect_presence(image, roi_df, lbph_model)
    threshold_data = set_threshold(pred)
    threshold_json = json.dumps(threshold_data)

    classroom = Classroom.query.filter_by(classroom_no=classroom_no).first()
    if classroom:
        classroom.seat_distribution = threshold_json
        db.session.commit()
        print(f'Seat distribution updated for classroom {classroom_no}')
    else:
        print(f'Classroom {classroom_no} not found')

    return jsonify({"status": "Training started"}), 202

@main.route('/training-status')
def training_status():
    # For simplicity, we just simulate the completion here
    return jsonify({"message": "Training has finished! You're all set to utilize the model."})

@main.route('/util/<classroom_no>')
def util(classroom_no):
    return render_template('util.html', classroom_no=classroom_no)

@main.route('/process_image', methods=['POST'])
def process_image():
    try:
        # Extract the JSON data from the request
        data = request.json
        
        # Validate required data
        if not data:
            return jsonify({'message': 'No data provided', 'success': False}), 400
        image_src = data.get('imageSrc')
        classroom_no = data.get('classroom_no')
        
        if not image_src or not classroom_no:
            return jsonify({'message': 'Missing imageSrc or classroom_no', 'success': False}), 400
        
        # Load saved LBPH model
        lbph_model_path = f'Classrooms/{classroom_no}/detector/trainingData.yml'
        if not os.path.exists(lbph_model_path):
            return jsonify({'message': 'Model file not found', 'success': False}), 404
        
        lbph_model = cv2.face.LBPHFaceRecognizer_create()
        lbph_model.read(lbph_model_path)
        
        # Load ROIs coordinates from roi.csv
        roi_csv_path = f'Classrooms/{classroom_no}/roi.csv'
        if not os.path.exists(roi_csv_path):
            return jsonify({'message': 'ROI file not found', 'success': False}), 404
        
        roi_df = pd.read_csv(roi_csv_path)
        
        # Read the image
        image = cv2.imread(image_src)
        if image is None:
            return jsonify({'message': 'Error reading image', 'success': False}), 400
        
        # Detect presence and mark attendance
        results = detect_presence(image, roi_df, lbph_model)
        mark_attendance(results, classroom_no)

        # Return a response indicating that the image processing was successful
        return jsonify({'message': 'Image processing complete', 'success': True})
    except Exception as e:
        # Handle any exceptions and return an error response
        return jsonify({'message': 'Error processing image: ' + str(e), 'success': False})


@main.route('/sheet/<classroom_no>')
def sheet(classroom_no):
    sheet_name = classroom_no
    html_table = extract_excel_to_html(sheet_name)
    # Pass the concatenated HTML table to the template for rendering
    return render_template('sheet.html', table=html_table, classroom_no=classroom_no)

@main.route('/generate_report/<classroom_no>')
def generate_report(classroom_no):
    html_table = extract_excel_to_html(classroom_no)
    pdf_content = create_attendance_report(html_table, classroom_no)
    return send_file(BytesIO(pdf_content), mimetype='application/pdf')

@main.route('/redirect_to_streamlit')
def redirect_to_streamlit():
    # Redirect to the URL where your Streamlit application is hosted
    return redirect("http://localhost:8501/", code=302)

app = create_app() # we initialize our flask app using the __init__.py function

if __name__ == '__main__':
    with app.app_context():
        db.create_all() # create the SQLite database
    app.run(debug=True) # run the flask app on debug mode
