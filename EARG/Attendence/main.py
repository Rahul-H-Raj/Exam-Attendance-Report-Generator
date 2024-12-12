import cv2
import numpy as np
import openpyxl
import pandas as pd
from PIL import Image
import streamlit as st

recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read('recognizer/trainingData.yml')

workbook = openpyxl.load_workbook('ExamSeatArrangement.xlsx')
sheet = workbook.active

face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

def mark_attendance(image):
    gray_image = cv2.cvtColor(image, cv2.COLOR_RGB2GRAY)

    faces = face_cascade.detectMultiScale(gray_image, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))

    attendance = []

    for (x, y, w, h) in faces:
        face = gray_image[y:y+h, x:x+w]

        id_, confidence = recognizer.predict(face)

        names = {1: "Seemanth", 2: "Prajwal", 3: "Venkatesh", 4: "Jithesh"}

        if confidence < 50:
            name = names.get(id_, "Unknown")
        else:
            name = "Unknown"
        attendance.append(name)

    for name in attendance:
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == name:
                sheet.cell(row=row, column=3).value = "Present"
                break

    workbook.save('ExamSeatArrangement.xlsx')

def capture_image():
    cap = cv2.VideoCapture(0)
    
    if not cap.isOpened():
        st.error("Error: Unable to open camera.")
        return
    
    cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
    
    while True:
        ret, frame = cap.read()
        
        if not ret:
            st.warning("Warning: Unable to capture frame from camera.")
            break
        
        cv2.imshow('Press Space to Capture', frame)
        
        if cv2.waitKey(1) & 0xFF == ord(' '):
            cv2.imwrite('captured_image.png', frame)
            break
    
    cap.release()
    cv2.destroyAllWindows()

# Streamlit app
def main():
    # Custom CSS styling
    st.markdown("""
    <style>
    body {
        background-color: #f0f2f6;
        font-family: Arial, sans-serif;
    }
    .title {
        color: #2c3e50;
        text-align: center;
    }
    .options {
        margin-top: 20px;
    }
    .button {
        background-color: #3498db;
        color: white;
        padding: 10px 20px;
        border: none;
        border-radius: 5px;
        cursor: pointer;
    }
    .button:hover {
        background-color: #2980b9;
    }
    </style>
    """, unsafe_allow_html=True)

    st.title("Customized Exam Attendance System")

    option = st.radio("Select an Option:", ("Upload an Image", "Capture an Image", "View Attendance Sheet"), key="option")

    if option == "Upload an Image":
        uploaded_file = st.file_uploader("Upload Image (JPG, JPEG, PNG)", type=["jpg", "jpeg", "png"])
        if uploaded_file is not None:
            image = Image.open(uploaded_file)
            st.image(image, caption='Uploaded Image', use_column_width=True)

            if st.button("Mark Attendance", key="mark_button"):
                mark_attendance(np.array(image))
                st.success("Attendance Marked Successfully!")

    elif option == "Capture an Image":
        st.write("Click the button below to capture an image from the camera and mark attendance")
        if st.button("Capture", key="capture_button"):
            st.write("Press the space bar to capture image")
            capture_image()
            captured_image = Image.open('captured_image.png')
            st.image(captured_image, caption='Captured Image', use_column_width=True)

            mark_attendance(np.array(captured_image))
            st.success("Attendance Marked Successfully!")

    elif option == "View Attendance Sheet":
        df = pd.read_excel('ExamSeatArrangement.xlsx')
        st.write(df)

if __name__ == "__main__":
    main()
