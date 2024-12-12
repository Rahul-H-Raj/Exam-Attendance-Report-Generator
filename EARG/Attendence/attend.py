import cv2
import numpy as np
import openpyxl

# Load the trained data
recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read('recognizer/trainingData.yml')

# Load the Excel workbook and worksheet
workbook = openpyxl.load_workbook('ExamSeatArrangement.xlsx')
sheet = workbook.active

# Load the image
image = cv2.imread('Image dataset/IMG_20240325_133933.jpg')

# Convert the image to grayscale
gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

# Load the face cascade
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

# Detect faces in the image
faces = face_cascade.detectMultiScale(gray_image, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))

# Initialize attendance list to keep track of marked attendees
attendance = []

# Iterate through detected faces
for (x, y, w, h) in faces:
    # Extract the face
    face = gray_image[y:y+h, x:x+w]

    # Recognize the face
    id_, confidence = recognizer.predict(face)

    names = {1: "Seemanth", 2: "Prajwal", 3: "Venkatesh", 4: "Jithesh"}

    if confidence < 50:
        name = names.get(id_, "Unknown")
    else:
        name = "Unknown"
    attendance.append(name)

# Mark attendance in the Excel sheet
for name in attendance:
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == name:
            sheet.cell(row=row, column=3).value = "Present"
            break

# Save the changes to the Excel file
workbook.save('ExamSeatArrangement.xlsx')