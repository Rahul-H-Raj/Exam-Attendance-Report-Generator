import cv2
import numpy as np
import pandas as pd
import openpyxl
# Load saved LBPH model
lbph_model = cv2.face.LBPHFaceRecognizer_create()
lbph_model.read('Classrooms/A 201/detector/trainingData.yml')

workbook = openpyxl.load_workbook('ExamSeatArrangement.xlsx')

# Function to predict on regions using LBPH model
def detect_presence(image, roi_df):
    predictions = {}
    for index, row in roi_df.iterrows():
        x, y, w, h = row['x'], row['y'], row['width'], row['height']
        # Extract region from the image
        region_img = image[y:y+h, x:x+w]
        # Convert region image to grayscale
        gray_region = cv2.cvtColor(region_img, cv2.COLOR_BGR2GRAY)
        # Make prediction using LBPH model
        label, confidence = lbph_model.predict(gray_region)
        predictions.update({index+1: (label, confidence)})
        print(predictions)
    return predictions

def mark_attendance(predictions, classroom_no):
    # Load the Excel workbook
    try:
        sheet = workbook[classroom_no]

        # Constants for threshold values
        THRESHOLDS = {
            1: 50, 2: 60, 3: 60, 4: 60, 
            5: 80, 6: 79, 7: 79, 8: 80, 
            9: 100, 10: 100, 11: 110, 12: 110
        }

        # Loop through rows
        for row in range(1, sheet.max_row + 1):
            if all(sheet.cell(row=row, column=col).value is None for col in range(1, 5)):
                break
            for col in range(1, sheet.max_column + 1, 4):
                sheet.cell(row=row, column=col + 3, value="Present")
                seat_no = sheet.cell(row=row, column=col).value
                if seat_no in predictions:
                    prediction = predictions[seat_no]
                    if prediction[1] < THRESHOLDS.get(seat_no, 0):
                        sheet.cell(row=row, column=col + 3, value="Absent")
    except FileNotFoundError:
        print("File not found.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")
    else:
        # Save the changes to the workbook
        workbook.save('ExamSeatArrangement.xlsx')
        # Close the workbook
        workbook.close()

# Load input image
img = cv2.imread('frame_3.jpg')

# Load ROIs coordinates from roi.csv
roi_df = pd.read_csv('roi.csv')

# Predict on regions
region_predictions = detect_presence(img, roi_df)
mark_attendance(region_predictions, 'A 201')